import openpyxl
from openpyxl import load_workbook
import tkinter as tk
from math import radians, sin, cos, sqrt, asin

filepath="/Users/michaelroth/Documents/School/12th Grade/APCSP/bank data.xlsx"
wb=load_workbook(filepath,data_only=True) # loads data from spreadsheet from filepath
sheet=wb.active

def distance(lat1, long1, lat2, long2):
    radius = 3958.8  # Earth radius in miles
 
    difLat = radians(lat2 - lat1) # difference between latitudes of coordinates
    difLong = radians(long2 - long1) # difference between longitudes of coordinates
    lat1 = radians(lat1)
    lat2 = radians(lat2)
 
    # below is haversine formula, which is used to determine the difference between two points on a sphere
    # the haversine formula closely follows the distances between coordinates on short distances, though Earth is not a perfect sphere
    step1 = sin(difLat / 2)**2 + cos(lat1) * cos(lat2) * sin(difLong / 2)**2
    step2 = 2 * asin(sqrt(step1))
        
    return (radius * step2)

# takes user's coordinates and mile radius to find every bank within the specified distance from the spreadsheet and returns the list of those banks
def findBank(lat1, long1, miles):
    banks = []

    for i in range(1, 87931): # iterates through the spreadsheet and calculates the distance from the user to the bank
        d = distance(lat1, long1, sheet.cell(i,3).value, sheet.cell(i,4).value)

        if(d < miles): # appends banks if bank is within specified distance to list
            banks.append(sheet.cell(i,1).value)
    
    return banks

# takes lists of every bank within distance and returns a list with the number of bank branches followed by the name of the bank
def countBanks(banks): 
    bankCount = []

    for i in range(len(banks)):
        if (banks[i] in banks[0:i]): # if bank name has been added previously the for loop continues to the next iteration
            continue
        
        bankCount.append(banks.count(banks[i])) # appends number of branches list
        bankCount.append(banks[i]) # appends name of bank to list
    
    return bankCount

# takes miles and list of number of branches and unique banks and updates the text widget to show the bank branches in the radius
def updateBankList(bankCount, miles):
    group = ""

    for i in range(0, len(bankCount), 2): # iterates through evens of bank count to create string of number of branches then bank
        if (bankCount[i] > 1): # determines if branches of bank are singular to use correct grammer when printing
            group = group + str(bankCount[i]) + " " + str(bankCount[i+1]) + " locations." + "\n"
        else:
            group = group + str(bankCount[i]) + " " + str(bankCount[i+1]) + " location." + "\n"

    text.delete('1.0', tk.END) # clears previous text in text widget for new bank search

    if (group != ""):
        text.insert('1.0', str("Banks in " + str(miles) + " mile radius: \n" + group)) # updates text widget with string of banks and number of branches

    else:
        text.insert('1.0', str("There are no banks within your " + str(miles) + " miles.")) # prints if no banks are found


HEIGHT=500
WIDTH=700

r = tk.Tk() # creates instance of gui

# creates title
r.title("Bank Finder")

# sets size of window using HEIGHT and WIDTH
canvas = tk.Canvas(r, height=HEIGHT, width=WIDTH)
canvas.pack()

# creates banner at the top
frame = tk.Frame(r, bg='#7fc6fb', bd=5)
frame.place(relx=0.5, rely=0.1, relwidth=0.85, relheight=0.1, anchor='n')

# creates button banner
frame2 = tk.Frame(r, bg='#7fc6fb', bd=5)
frame2.place(relx=0.7735, rely=0.2, relwidth=0.3, relheight=0.1, anchor='n')

# creates leftmost label
label1 = tk.Label(frame, text="x: ")
label1.place(relwidth=0.1, relheight=1)

# creates middle label
label2 = tk.Label(frame, text="y: ")
label2.place(relx=.34, relwidth=0.1, relheight=1)

# creates rightmost label
label3 = tk.Label(frame, text="Radius: ")
label3.place(relx=.685, relwidth=0.1, relheight=1)

# creates leftmost entry box
entry1 = tk.Entry(frame, font=100)
entry1.place(relx=.13,relwidth=0.2, relheight=1)

# creates middle entry box
entry2 = tk.Entry(frame, font=100)
entry2.place(relx=.47, relwidth=0.20, relheight=1)

# creates rightmost entry box
entry3 = tk.Entry(frame, font=100)
entry3.place(relx=.8, relwidth=0.20, relheight=1)

# creates textbox to display banks
text=tk.Text(r,font=100)
text.place(relx=0.5, rely=0.32, relwidth=0.85, relheight=0.65, anchor='n')

# creates button to calculate results
button = tk.Button(frame2, text="Find Banks", font=40, command=lambda: updateBankList(countBanks(findBank(float(entry1.get()), float(entry2.get()),int(entry3.get()))), int(entry3.get()))) # gets entry from boxes to input into functions
button.place(relheight=1, relwidth=1)
r.mainloop()